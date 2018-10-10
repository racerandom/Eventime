# coding=utf-8

import openpyxl
import os
from collections import defaultdict
from krippendorff_alpha import *


ts_dir = "/Users/fei-c/Resources/timex/Release0531/比較用/ts"
kd_dir = "/Users/fei-c/Resources/timex/Release0531/比較用/kd"




def read_event_dic(file_dir):

    event_dic = defaultdict(dict)

    for filename in sorted(os.listdir(file_dir)):
        file = os.path.join(file_dir, filename)
        try:
            wb = openpyxl.load_workbook(file)
            first_sheet = wb.sheetnames[0]
            worksheet = wb[first_sheet]

            #here you iterate over the rows in the specific column
            for row in range(7,worksheet.max_row+1):
                eid = worksheet["E{}".format(row)].value
                if eid:
                    event_dic[filename][eid] = worksheet["H{}".format(row)].value
        except Exception as ex:
            print("Error in Reading file, ", file)

    return event_dic


ts_dic = read_event_dic(ts_dir)
kd_dic = read_event_dic(kd_dir)
# print(ts_dic)

total, all, correct, scorrect, type_diff, type_same, single, multi, psedo_correct, none_case = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
ts_out, kd_out = [], []

def diff(a, b):
    if "beginPoint" in a and "beginPoint" not in b:
        return True
    elif "beginPoint" in b and "beginPoint" not in a:
        return True
    else:
        return False

def isMulti(a):
    if "beginPoint" in a:
        return True
    else:
        return False


for file in kd_dic.keys():
    ts_events = ts_dic[file]
    kd_events = kd_dic[file]

    common_event_ids = list(set(ts_events.keys()) & set(kd_events.keys()))

    for event_id in common_event_ids:
        ts_anchor = ts_events[event_id]
        kd_anchor = kd_events[event_id]
        ts_out.append(' '.join(ts_anchor.split()) if ts_anchor else None)
        kd_out.append(' '.join(kd_anchor.split()) if kd_anchor else None)
        if not ts_anchor or not kd_anchor:
            none_case += 1


    for event, kd_anchor in kd_events.items():
        total += 1
        if event in ts_events:
            ts_anchor = ts_events[event]
            if ts_anchor and kd_anchor:
                all += 1
                # ts_out.append(1)
                ts_norm = ''.join(ts_anchor.split())
                kd_norm = ''.join(kd_anchor.split())

                if ts_norm == kd_norm:
                    # print(ts_norm, kd_norm)
                    if not isMulti(ts_norm):
                        scorrect += 1
                    correct += 1
                    # ks_out.append(1)
                else:
                    print(ts_norm.replace('beginPoint=', '').replace('endPoint=', ''), kd_norm.replace('beginPoint=', '').replace('endPoint=', ''))
                    if ts_norm.replace('beginPoint=', '').replace('endPoint=', '') == kd_norm.replace('beginPoint=', '').replace('endPoint=', ''):
                        psedo_correct += 1
                    if diff(ts_norm, kd_norm):
                        type_diff += 1
                    else:
                        if isMulti(ts_norm):
                            multi +=1
                        else:
                            single += 1
                        type_same += 1
                    # ks_out.append(2)
                # print(event, ''.join(ts_anchor.split()), ''.join(kd_anchor.split()))

print(total, all, correct, scorrect, type_diff, type_same, "diff", multi, single, psedo_correct, none_case)
print(correct / all)
print("nominal metric: %.3f" % krippendorff_alpha([ts_out, kd_out], nominal_metric, convert_items=str, missing_items=None))
print("interval metric: %.3f" % krippendorff_alpha([ts_out, kd_out], nominal_metric, convert_items=str, missing_items=None))


